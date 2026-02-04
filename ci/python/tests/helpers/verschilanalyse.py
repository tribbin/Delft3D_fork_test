import random
from datetime import datetime, timedelta, timezone

import openpyxl

from ci_tools.verschilanalyse.util.slurm_log_data import SlurmLogData
from ci_tools.verschilanalyse.util.verschilanalyse_comparison import VerschilanalyseComparison
from ci_tools.verschilanalyse.util.verschillentool import (
    OutputType,
    Statistics,
    Tolerances,
    Variable,
    VerschillentoolOutput,
)


def make_log_data(
    commit_id: str | None = None,
    slurm_job_id: str = "",
    slurm_nodelist: str = "",
    task_count: int = 8,
    mean_computation_time: float | None = None,
    stddev_computation_time: float | None = None,
    warning_lines: list[tuple[int, str]] | None = None,
    error_lines: list[tuple[int, str]] | None = None,
) -> SlurmLogData:
    """Test helper factory for `SlurmLogData` instances.

    Has default values for all parameters to make it easier to create test instances.
    """
    commit_id = commit_id or "".join(random.choices("0123456789abcdef", k=32))
    slurm_job_id = slurm_job_id or str(random.randint(int(1e5), int(1e6 - 1)))
    slurm_nodelist = slurm_nodelist or f"v-100vcpu-{random.randint(10, 99)}"
    if mean_computation_time is None:
        mean_computation_time = random.random() * 86400.0
    if stddev_computation_time is None:
        stddev_computation_time = random.random() * 42.0
    warning_lines = warning_lines or []
    error_lines = error_lines or []

    return SlurmLogData(
        commit_id=commit_id,
        slurm_job_id=slurm_job_id,
        slurm_nodelist=slurm_nodelist,
        task_count=task_count,
        mean_computation_time=mean_computation_time,
        stddev_computation_time=stddev_computation_time,
        warning_lines=warning_lines,
        error_lines=error_lines,
    )


def make_verschillentool_output(
    output_type: OutputType = OutputType.HIS,
    water_level: Statistics | None = None,
    flow_velocity: Statistics | None = None,
    row_count: int = 7,
) -> VerschillentoolOutput:
    """Test helper factory for `VerschillentoolOuput` instances.

    Has default values for all parameters to make it easier to create test instances.
    """
    flow_velocity = flow_velocity or Statistics(0.0, 0.0, 0.0, 0.0)
    water_level = water_level or Statistics(0.0, 0.0, 0.0, 0.0)
    return VerschillentoolOutput(
        output_type=output_type,
        water_level=water_level,
        flow_velocity=flow_velocity,
        row_count=row_count,
    )


def make_verschilanalyse_comparison(
    s3_current_prefix: str = "s3://bucket/output/current",
    s3_reference_prefix: str = "s3://bucket/output/reference",
    current_log_data: dict[str, SlurmLogData] | None = None,
    reference_log_data: dict[str, SlurmLogData] | None = None,
    his_outputs: dict[str, VerschillentoolOutput] | None = None,
    map_outputs: dict[str, VerschillentoolOutput] | None = None,
) -> VerschilanalyseComparison:
    """Test helper actory for `VerschilanalyseComparison` instances.

    Has default values for all parameters to make it easier to create test instances.
    """
    default_models = ["foo", "bar", "baz"]
    if current_log_data is None:
        current_log_data = {name: make_log_data() for name in default_models}
    if reference_log_data is None:
        reference_log_data = {name: make_log_data() for name in default_models}
    if his_outputs is None:
        his_outputs = {name: make_verschillentool_output(output_type=OutputType.HIS) for name in default_models}
    if map_outputs is None:
        map_outputs = {name: make_verschillentool_output(output_type=OutputType.MAP) for name in default_models}

    return VerschilanalyseComparison(
        s3_current_prefix=s3_current_prefix,
        s3_reference_prefix=s3_reference_prefix,
        current_log_data=current_log_data,
        reference_log_data=reference_log_data,
        his_outputs=his_outputs,
        map_outputs=map_outputs,
    )


def make_verschillentool_workbook(
    water_level_stats: Statistics | None = None,
    flow_velocity_stats: Statistics | None = None,
    row_count: int = 10,
    output_type: OutputType = OutputType.HIS,
) -> openpyxl.Workbook:
    if water_level_stats is None:
        water_level_stats = Statistics(0.0, 0.0, 0.0, 0.0)
    if flow_velocity_stats is None:
        flow_velocity_stats = Statistics(0.0, 0.0, 0.0, 0.0)

    workbook = openpyxl.Workbook()

    # Write statistics sheet.
    stats_sheet = workbook.active
    if stats_sheet is None:
        raise RuntimeError("active sheet is None")
    stats_sheet.title = "Statistics"
    if output_type == OutputType.HIS:
        stats_sheet.append(["stations", "station_x_coordinate", "station_y_coordinate"])
        for i in range(row_count):
            stats_sheet.append([f"station_{i}", random.random() * 100, random.random() * 100])
    else:
        now = datetime.now(timezone.utc) - timedelta(days=row_count)
        stats_sheet.append(["time", "sea_surface_height_bias", "sea_water_level_bias"])
        for i in range(row_count):
            stats_sheet.append([(now + timedelta(days=i)).isoformat(), random.random(), random.random()])

    # Leave maxima sheet empty for now.
    maxima_sheet = workbook.create_sheet(title="Maxima")
    for row in [
        ["", "Time", "Maximum value over all times", "layers of max"],
        ["sea_surface_height (m)", "2-1-2035  11:00:00", water_level_stats.max, "2D variable"],
        ["sea_water_speed (m s-1)", "2-1-2035  06:00:00", flow_velocity_stats.max, "2D variable"],
    ]:
        maxima_sheet.append(row)

    # Write averages sheet.
    averages_sheet = workbook.create_sheet(title="Averages")
    for row in [
        ["", "Average over all stations"],
        ["sea_water_speed_max (m s-1)", flow_velocity_stats.avg_max],
        ["sea_water_speed_bias (m s-1)", flow_velocity_stats.avg_bias],
        ["sea_water_speed_rms (m s-1)", flow_velocity_stats.avg_rms],
        ["sea_surface_height_max (m)", water_level_stats.avg_max],
        ["sea_surface_height_bias (m)", water_level_stats.avg_bias],
        ["sea_surface_height_rms (m)", water_level_stats.avg_rms],
    ]:
        averages_sheet.append(row)

    return workbook


def tolerance_stats(output_type: OutputType, variable: Variable, diff: float = 0.0) -> Statistics:
    """Set statistics to the maximum allowed tolerances and optionally add a difference value.

    The maximum allowed tolerance values depend on the `OutputType` and `Variable`.

    Parameters
    ----------
    output_type : OutputType
    variable : Variable
    diff : float, optional
        Add this value to the maximum tolerance values. Pass a
        negative value here to stay under the tolerance threshold.

    Returns
    -------
    Statistics
    """
    return Statistics(
        avg_max=Tolerances.max(output_type, variable) + diff,
        avg_bias=Tolerances.bias(output_type, variable) + diff,
        avg_rms=Tolerances.rms(output_type, variable) + diff,
        max=Tolerances.max(output_type, variable) + diff,
    )
