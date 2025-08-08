"""Tests for update_excel_sheet.py."""

from datetime import datetime, timezone
from unittest.mock import MagicMock, Mock, call, patch

import pytest
from openpyxl.worksheet.worksheet import Worksheet

from ci_tools.dimrset_delivery.common_utils import ResultTestBankParser
from ci_tools.dimrset_delivery.dimr_context import DimrAutomationContext
from ci_tools.dimrset_delivery.lib.teamcity import TeamCity
from ci_tools.dimrset_delivery.settings.general_settings import DRY_RUN_PREFIX, VERSIONS_EXCEL_FILENAME
from ci_tools.dimrset_delivery.update_excel_sheet import ExcelHelper, update_excel_sheet


class TestExcelHelper:
    """Test cases for the ExcelHelper class."""

    def setup_method(self) -> None:
        """Set up test fixtures before each test method."""
        # Arrange - Common test data
        self.mock_teamcity = Mock(spec=TeamCity)
        self.filepath = "test_versions.xlsx"
        self.dimr_version = "v1.2.3"
        self.kernel_versions = {"build.vcs.number": "12345", "other_kernel": "v2.0.0"}
        self.mock_parser = Mock(spec=ResultTestBankParser)

        # Configure mock parser return values
        self.mock_parser.get_percentage_total_passing.return_value = "85.5%"
        self.mock_parser.get_total_tests.return_value = 100
        self.mock_parser.get_total_passing.return_value = 85
        self.mock_parser.get_total_failing.return_value = 10
        self.mock_parser.get_total_exceptions.return_value = 5

    def create_excel_helper(self) -> ExcelHelper:
        """Create an ExcelHelper instance with test data."""
        return ExcelHelper(
            teamcity=self.mock_teamcity,
            filepath=self.filepath,
            dimr_version=self.dimr_version,
            kernel_versions=self.kernel_versions,
            parser=self.mock_parser,
        )

    @patch("ci_tools.dimrset_delivery.update_excel_sheet.load_workbook")
    @patch("ci_tools.dimrset_delivery.update_excel_sheet.datetime")
    def test_append_row_successfully_adds_new_row(self, mock_datetime: Mock, mock_load_workbook: Mock) -> None:
        """Test that append_row successfully adds a new row when conditions are met."""
        # Arrange
        helper = self.create_excel_helper()

        # Mock datetime to return a fixed date
        mock_date = datetime(2025, 8, 8, tzinfo=timezone.utc)
        mock_datetime.now.return_value = mock_date

        # Mock workbook and worksheet
        mock_workbook = MagicMock()
        mock_worksheet = MagicMock(spec=Worksheet)
        mock_load_workbook.return_value = mock_workbook

        # Configure the workbook to return the worksheet when accessed with SHEET_NAME
        mock_workbook.__getitem__.return_value = mock_worksheet

        # Mock worksheet does not contain the row already
        mock_worksheet.__getitem__.return_value = []  # Empty name column

        expected_row = [
            "",  # Column A (empty)
            "2025-08-08",  # Column B (Date)
            "DIMRset v1.2.3",  # Column C (DIMR version)
            "",  # Column D (Revision)
            "FLOW1D2D now in GitHub",  # Column E (Flow1D)
            "OSS",  # Column F (FlowFM)
            "12345",  # Column G (OSS)
            "DRR now in GitHub",  # Column H (RR)
            "FBC now in GitHub",  # Column I (FBC)
            "85.5%",  # Column J (Percentage passing)
            100,  # Column K (Total tests)
            85,  # Column L (Passing tests)
            10,  # Column M (Failing tests)
            5,  # Column N (Exception tests)
            "",  # Column O (Docker hub)
            "Flow1D and RR: only Windows",  # Column P (Remarks)
        ]

        # Act
        helper.append_row()

        # Assert
        mock_load_workbook.assert_called_once_with(filename=self.filepath)
        mock_worksheet.append.assert_called_once_with(expected_row)
        mock_workbook.save.assert_called_once_with(filename=self.filepath)
        mock_workbook.close.assert_called_once()

    @patch("ci_tools.dimrset_delivery.update_excel_sheet.load_workbook")
    def test_append_row_skips_when_row_already_exists(self, mock_load_workbook: Mock) -> None:
        """Test that append_row skips adding row when it already exists."""
        # Arrange
        helper = self.create_excel_helper()

        # Mock workbook and worksheet
        mock_workbook = MagicMock()
        mock_worksheet = MagicMock(spec=Worksheet)
        mock_load_workbook.return_value = mock_workbook
        mock_workbook.__getitem__.return_value = mock_worksheet

        # Mock worksheet already contains the row
        mock_cell = Mock()
        mock_cell.value = f"DIMRset {self.dimr_version}"
        mock_worksheet.__getitem__.return_value = [mock_cell]

        # Act
        helper.append_row()

        # Assert
        mock_load_workbook.assert_called_once_with(filename=self.filepath)
        mock_worksheet.append.assert_not_called()  # Should not append
        mock_workbook.save.assert_not_called()  # Should not save
        mock_workbook.close.assert_called_once()

    @patch("ci_tools.dimrset_delivery.update_excel_sheet.load_workbook")
    @patch("builtins.print")
    def test_append_row_handles_exception_gracefully(self, mock_print: Mock, mock_load_workbook: Mock) -> None:
        """Test that append_row handles exceptions gracefully."""
        # Arrange
        helper = self.create_excel_helper()

        # Mock workbook to raise exception
        mock_workbook = MagicMock()
        mock_load_workbook.return_value = mock_workbook
        mock_workbook.__getitem__.side_effect = Exception("Test exception")

        # Act
        helper.append_row()

        # Assert
        mock_load_workbook.assert_called_once_with(filename=self.filepath)
        mock_workbook.close.assert_called_once()
        # Check that error messages were printed
        mock_print.assert_any_call("Could not update the excel: \n")

    @patch("ci_tools.dimrset_delivery.update_excel_sheet.datetime")
    def test_prepare_row_to_insert_returns_correct_format(self, mock_datetime: Mock) -> None:
        """Test that _prepare_row_to_insert returns the correctly formatted row."""
        # Arrange
        helper = self.create_excel_helper()

        # Mock datetime to return a fixed date
        mock_date = datetime(2025, 8, 8, tzinfo=timezone.utc)
        mock_datetime.now.return_value = mock_date

        # Act
        result = helper._ExcelHelper__prepare_row_to_insert()

        # Assert
        expected_row = [
            "",  # Column A (empty)
            "2025-08-08",  # Column B (Date)
            "DIMRset v1.2.3",  # Column C (DIMR version)
            "",  # Column D (Revision)
            "FLOW1D2D now in GitHub",  # Column E (Flow1D)
            "OSS",  # Column F (FlowFM)
            "12345",  # Column G (OSS)
            "DRR now in GitHub",  # Column H (RR)
            "FBC now in GitHub",  # Column I (FBC)
            "85.5%",  # Column J (Percentage passing)
            100,  # Column K (Total tests)
            85,  # Column L (Passing tests)
            10,  # Column M (Failing tests)
            5,  # Column N (Exception tests)
            "",  # Column O (Docker hub)
            "Flow1D and RR: only Windows",  # Column P (Remarks)
        ]

        assert result == expected_row

        # Verify parser methods were called
        self.mock_parser.get_percentage_total_passing.assert_called_once()
        self.mock_parser.get_total_tests.assert_called_once()
        self.mock_parser.get_total_passing.assert_called_once()
        self.mock_parser.get_total_failing.assert_called_once()
        self.mock_parser.get_total_exceptions.assert_called_once()

    def test_worksheet_already_contains_row_returns_true_when_exists(self) -> None:
        """Test that _worksheet_already_contains_row returns True when row exists."""
        # Arrange
        helper = self.create_excel_helper()
        mock_worksheet = MagicMock(spec=Worksheet)

        # Mock cells in name column
        mock_cell1 = Mock()
        mock_cell1.value = "DIMRset v1.0.0"
        mock_cell2 = Mock()
        mock_cell2.value = f"DIMRset {self.dimr_version}"  # This should match
        mock_cell3 = Mock()
        mock_cell3.value = "DIMRset v2.0.0"

        # Configure mock to return the list when accessing the column
        mock_worksheet.__getitem__.return_value = [mock_cell1, mock_cell2, mock_cell3]

        # Act
        result = helper._ExcelHelper__worksheet_already_contains_row(mock_worksheet)

        # Assert
        assert result is True

    def test_worksheet_already_contains_row_returns_false_when_not_exists(self) -> None:
        """Test that _worksheet_already_contains_row returns False when row doesn't exist."""
        # Arrange
        helper = self.create_excel_helper()
        mock_worksheet = MagicMock(spec=Worksheet)

        # Mock cells in name column that don't match
        mock_cell1 = Mock()
        mock_cell1.value = "DIMRset v1.0.0"
        mock_cell2 = Mock()
        mock_cell2.value = "DIMRset v2.0.0"

        # Configure mock to return the list when accessing the column
        mock_worksheet.__getitem__.return_value = [mock_cell1, mock_cell2]

        # Act
        result = helper._ExcelHelper__worksheet_already_contains_row(mock_worksheet)

        # Assert
        assert result is False

    def test_worksheet_already_contains_row_handles_empty_column(self) -> None:
        """Test that _worksheet_already_contains_row handles empty name column."""
        # Arrange
        helper = self.create_excel_helper()
        mock_worksheet = MagicMock(spec=Worksheet)

        # Mock empty name column
        mock_worksheet.__getitem__.return_value = []

        # Act
        result = helper._ExcelHelper__worksheet_already_contains_row(mock_worksheet)

        # Assert
        assert result is False


class TestUpdateExcelSheet:
    """Test cases for the update_excel_sheet function."""

    def setup_method(self) -> None:
        """Set up test fixtures before each test method."""
        # Arrange - Common test data
        self.mock_context = Mock(spec=DimrAutomationContext)
        self.mock_context.dry_run = False
        self.mock_context.ssh_client = Mock()
        self.mock_context.teamcity = Mock(spec=TeamCity)

        # Configure context methods
        self.kernel_versions = {"build.vcs.number": "12345"}
        self.dimr_version = "v1.2.3"
        self.mock_context.get_kernel_versions.return_value = self.kernel_versions
        self.mock_context.get_dimr_version.return_value = self.dimr_version

    @patch("ci_tools.dimrset_delivery.update_excel_sheet.get_testbank_result_parser")
    @patch("ci_tools.dimrset_delivery.update_excel_sheet.ExcelHelper")
    @patch("builtins.print")
    def test_update_excel_sheet_successful_execution(
        self, mock_print: Mock, mock_excel_helper_class: Mock, mock_get_parser: Mock
    ) -> None:
        """Test successful execution of update_excel_sheet function."""
        # Arrange
        mock_parser = Mock(spec=ResultTestBankParser)
        mock_get_parser.return_value = mock_parser

        mock_excel_helper = Mock()
        mock_excel_helper_class.return_value = mock_excel_helper

        # Act
        update_excel_sheet(self.mock_context)

        # Assert
        self.mock_context.print_status.assert_called_once_with("Updating Excel sheet...")
        self.mock_context.get_kernel_versions.assert_called_once()
        self.mock_context.get_dimr_version.assert_called_once()

        # Verify SSH operations
        assert self.mock_context.ssh_client.secure_copy.call_count == 2

        # Verify ExcelHelper creation and usage
        mock_excel_helper_class.assert_called_once_with(
            teamcity=self.mock_context.teamcity,
            filepath=VERSIONS_EXCEL_FILENAME,
            dimr_version=self.dimr_version,
            kernel_versions=self.kernel_versions,
            parser=mock_parser,
        )
        mock_excel_helper.append_row.assert_called_once()

        # Verify success message
        mock_print.assert_called_with("Excel sheet update completed successfully!")

    @patch("builtins.print")
    def test_update_excel_sheet_dry_run_mode(self, mock_print: Mock) -> None:
        """Test update_excel_sheet function in dry run mode."""
        # Arrange
        self.mock_context.dry_run = True

        # Act
        update_excel_sheet(self.mock_context)

        # Assert
        self.mock_context.print_status.assert_called_once_with("Updating Excel sheet...")
        self.mock_context.get_kernel_versions.assert_called_once()
        self.mock_context.get_dimr_version.assert_called_once()

        # Verify dry run messages
        expected_calls = [
            call(f"{DRY_RUN_PREFIX} Would update Excel sheet with DIMR version:", self.dimr_version),
            call(f"{DRY_RUN_PREFIX} Would download Excel from network drive"),
            call(f"{DRY_RUN_PREFIX} Would append new row with release information"),
            call(f"{DRY_RUN_PREFIX} Would upload updated Excel back to network drive"),
        ]
        mock_print.assert_has_calls(expected_calls)

        # Verify no actual operations were performed
        assert (
            not hasattr(self.mock_context.ssh_client, "secure_copy")
            or self.mock_context.ssh_client.secure_copy.call_count == 0
        )

    @patch("ci_tools.dimrset_delivery.update_excel_sheet.get_testbank_result_parser")
    def test_update_excel_sheet_raises_error_when_ssh_client_missing(self, mock_get_parser: Mock) -> None:
        """Test that update_excel_sheet raises ValueError when SSH client is missing."""
        # Arrange
        mock_parser = Mock(spec=ResultTestBankParser)
        mock_get_parser.return_value = mock_parser
        self.mock_context.ssh_client = None

        # Act & Assert
        with pytest.raises(ValueError, match="SSH client is required but not initialized"):
            update_excel_sheet(self.mock_context)

    @patch("ci_tools.dimrset_delivery.update_excel_sheet.get_testbank_result_parser")
    def test_update_excel_sheet_raises_error_when_teamcity_missing(self, mock_get_parser: Mock) -> None:
        """Test that update_excel_sheet raises ValueError when TeamCity client is missing."""
        # Arrange
        mock_parser = Mock(spec=ResultTestBankParser)
        mock_get_parser.return_value = mock_parser
        self.mock_context.teamcity = None

        # Act & Assert
        with pytest.raises(ValueError, match="TeamCity client is required but not initialized"):
            update_excel_sheet(self.mock_context)

    @patch("ci_tools.dimrset_delivery.update_excel_sheet.get_testbank_result_parser")
    def test_update_excel_sheet_calls_get_testbank_result_parser(self, mock_get_parser: Mock) -> None:
        """Test that update_excel_sheet calls get_testbank_result_parser when not in dry run."""
        # Arrange
        mock_parser = Mock(spec=ResultTestBankParser)
        mock_get_parser.return_value = mock_parser

        # Act
        with patch("ci_tools.dimrset_delivery.update_excel_sheet.ExcelHelper"):
            update_excel_sheet(self.mock_context)

        # Assert
        mock_get_parser.assert_called_once()

    @patch("ci_tools.dimrset_delivery.update_excel_sheet.get_testbank_result_parser")
    def test_update_excel_sheet_does_not_call_parser_in_dry_run(self, mock_get_parser: Mock) -> None:
        """Test that update_excel_sheet doesn't call get_testbank_result_parser in dry run mode."""
        # Arrange
        self.mock_context.dry_run = True

        # Act
        update_excel_sheet(self.mock_context)

        # Assert
        mock_get_parser.assert_not_called()
