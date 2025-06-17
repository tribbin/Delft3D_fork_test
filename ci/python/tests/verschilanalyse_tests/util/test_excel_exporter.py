import pytest
from openpyxl.styles import PatternFill

from ci_tools.verschilanalyse.util.excel_exporter import ExcelExporter
from ci_tools.verschilanalyse.util.verschillentool import OutputType, Variable
from tests.helpers import verschilanalyse as helper


def test_make_summary_workbook__check_sheetnames() -> None:
    verschilanalyse_comparison = helper.make_verschilanalyse_comparison()

    workbook = ExcelExporter.make_summary_workbook(verschilanalyse_comparison)

    assert workbook.sheetnames == [
        ExcelExporter.LOG_COMPARISON_SHEET_NAME,
        ExcelExporter.VERSCHILLENTOOL_SHEET_NAMES[OutputType.HIS],
        ExcelExporter.VERSCHILLENTOOL_SHEET_NAMES[OutputType.MAP],
    ]


def test_make_summary_workbook__log_comparisons() -> None:
    # Arrange
    current_log_data = {"foo": helper.make_log_data(), "bar": helper.make_log_data()}
    reference_log_data = {"foo": helper.make_log_data(), "baz": helper.make_log_data()}
    verschilanalyse_comparison = helper.make_verschilanalyse_comparison(
        s3_current_prefix="s3://bucket/output/current",
        s3_reference_prefix="s3://bucket/output/reference",
        current_log_data=current_log_data,
        reference_log_data=reference_log_data,
    )

    # Act
    workbook = ExcelExporter.make_summary_workbook(verschilanalyse_comparison)
    log_comp_sheet = workbook[ExcelExporter.LOG_COMPARISON_SHEET_NAME]

    # Assert output and report paths are set correctly.
    assert [cell[0].value for cell in log_comp_sheet["B1:B4"]] == [
        "s3://bucket/output/current/output",
        "s3://bucket/output/current/logs/logs.zip",
        "s3://bucket/output/reference/output",
        "s3://bucket/output/reference/logs/logs.zip",
    ]

    # bar comparison is first (sorted by name).
    bar_header = log_comp_sheet["A6:C6"][0]
    assert bar_header[0].value == "⚠️ bar"  # bar is in current but not in reference.
    assert bar_header[1].value == "✅ Current"
    assert bar_header[2].value == "⚠️ Reference"

    # foo comparison is second.
    foo_header = log_comp_sheet["A17:C17"][0]
    assert foo_header[0].value == "✅ foo"  #  foo is in both current and reference.
    assert foo_header[1].value == "✅ Current"
    assert foo_header[2].value == "✅ Reference"


@pytest.mark.parametrize("output_type", OutputType)
def test_make_summary_workbook(output_type: OutputType) -> None:
    # Arrange
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    foo_output = helper.make_verschillentool_output(  # All stats within tolerance.
        output_type=output_type,
        water_level=helper.tolerance_stats(output_type, Variable.WATER_LEVEL, diff=-1e-6),
        flow_velocity=helper.tolerance_stats(output_type, Variable.FLOW_VELOCITY, diff=-1e-6),
        row_count=42,
    )
    bar_output = helper.make_verschillentool_output(  # All stats over tolerance.
        output_type=output_type,
        water_level=helper.tolerance_stats(output_type, Variable.WATER_LEVEL, diff=1e-6),
        flow_velocity=helper.tolerance_stats(output_type, Variable.FLOW_VELOCITY, diff=1e-6),
        row_count=43,
    )

    outputs = {"foo": foo_output, "bar": bar_output}
    count_header = ExcelExporter.VERSCHILLENTOOL_COUNT_HEADERS[output_type]
    sheet_title = ExcelExporter.VERSCHILLENTOOL_SHEET_NAMES[output_type]
    if output_type == OutputType.HIS:
        verschilanalyse_comparison = helper.make_verschilanalyse_comparison(his_outputs=outputs)
    else:  # output_type == OutputType.MAP
        verschilanalyse_comparison = helper.make_verschilanalyse_comparison(map_outputs=outputs)

    # Act
    workbook = ExcelExporter.make_summary_workbook(verschilanalyse_comparison)
    sheet = workbook[sheet_title]
    header_row, bar_row, foo_row = sheet["A1:H3"]

    # Assert
    assert header_row[1].value == count_header

    # Assert all bar statistics are over tolerance.
    assert bar_row[0].value == "bar"
    assert bar_row[1].value == 43
    assert all(cell.fill == red_fill and str(cell.value).startswith("❌") for cell in bar_row[2:5])
    assert all(cell.fill == red_fill and str(cell.value).startswith("❌") for cell in bar_row[6:9])

    # Assert all foo statistics are within tolerance.
    assert foo_row[0].value == "foo"
    assert foo_row[1].value == 42
    assert all(cell.fill != red_fill and not str(cell.value).startswith("❌") for cell in foo_row[2:])
