import argparse
from dataclasses import dataclass
from enum import Enum
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill

MAX_TOLERANCE = 0.01
RMS_TOLERANCE = 0.001
FLOWVELOCITY_MEAN_TOLERANCE = 0.0005
WATERLEVEL_MEAN_TOLERANCE = 0.0001


class FileType(Enum):
    """Type of output netcdf file."""

    HIS = "his"
    MAP = "map"


@dataclass
class Statistics:
    """Statistics of a sample."""

    max: float
    mean: float
    rms: float


@dataclass
class ModelStatistics:
    """Contains statistics of a model run."""

    flow_velocity: Statistics
    water_level: Statistics
    row_count: int


def parse_dataframes(root_dir: Path, file_type: FileType) -> dict[str, ModelStatistics]:
    """Read verschillentool excel files."""
    result: dict[str, ModelStatistics] = {}

    file_name = f"{file_type.value}_output.xlsx"
    for model_dir in (path for path in root_dir.iterdir() if (path / file_name).exists()):
        file_path = model_dir / file_name
        averages_sheet = pd.read_excel(file_path, sheet_name="Averages")
        statistics_sheet = pd.read_excel(file_path, sheet_name="Statistics")

        # For some reason, the flow velocity and water level data are swapped in the different output files.
        if file_type == FileType.HIS:
            column = averages_sheet["Average over all stations"].round(4)
            flow_velocity_stats = Statistics(max=column[0], mean=column[1], rms=column[2])
            water_level_stats = Statistics(max=column[3], mean=column[4], rms=column[5])
        elif file_type == FileType.MAP:
            column = averages_sheet["Average over all times"].round(4)
            water_level_stats = Statistics(max=column[0], mean=column[1], rms=column[2])
            flow_velocity_stats = Statistics(max=column[3], mean=column[4], rms=column[5])
        else:
            raise ValueError(f"File type '{file_type}' not supported")

        result[model_dir.name] = ModelStatistics(
            flow_velocity=flow_velocity_stats,
            water_level=water_level_stats,
            row_count=len(statistics_sheet),
        )

    return result


def write_excel_water_level(
    output_dir: Path,
    stats_per_model: dict[str, ModelStatistics],
    file_type: FileType,
) -> None:
    """Write dataframes to excel file."""
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise ValueError("Active work sheet is 'None'")

    row_count_name = "#stations" if file_type == FileType.HIS else "#maps"

    red_solid_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    ws.append(["Naam Model", row_count_name, "Maximum", "Bias", "RMSE"])
    for model_name, model_stats in sorted(stats_per_model.items()):
        stats = model_stats.water_level
        station_count = model_stats.row_count

        ws.append([model_name, station_count, stats.max, stats.mean, stats.rms])
        if stats.max > MAX_TOLERANCE:
            ws.cell(row=ws.max_row, column=3).fill = red_solid_fill
        if stats.mean > WATERLEVEL_MEAN_TOLERANCE:
            ws.cell(row=ws.max_row, column=4).fill = red_solid_fill
        if stats.rms > RMS_TOLERANCE:
            ws.cell(row=ws.max_row, column=5).fill = red_solid_fill

    wb.save(str(output_dir / f"{file_type.value}_waterlevel.xlsx"))


def write_excel_flow_velocity(
    output_dir: Path,
    stats_per_model: dict[str, ModelStatistics],
    file_type: FileType,
) -> None:
    """Write dataframes to excel file."""
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise ValueError("Active work sheet is 'None'")

    row_count_name = "#stations" if file_type == FileType.HIS else "#maps"

    red_solid_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    ws.append(["Naam Model", row_count_name, "Maximum", "Bias", "RMSE"])
    for model_name, model_stats in sorted(stats_per_model.items()):
        stats = model_stats.flow_velocity
        station_count = model_stats.row_count

        ws.append([model_name, station_count, stats.max, stats.mean, stats.rms])
        if stats.max > MAX_TOLERANCE:
            ws.cell(row=ws.max_row, column=3).fill = red_solid_fill
        if stats.mean > FLOWVELOCITY_MEAN_TOLERANCE:
            ws.cell(row=ws.max_row, column=4).fill = red_solid_fill
        if stats.rms > RMS_TOLERANCE:
            ws.cell(row=ws.max_row, column=5).fill = red_solid_fill

    wb.save(str(output_dir / f"{file_type.value}_flowvelocity.xlsx"))


def make_argument_parser() -> argparse.ArgumentParser:
    """Make command line argument parser."""
    parser = argparse.ArgumentParser(description="Summary the verschillentool output in a few excel files.")
    parser.add_argument(
        "--verschillentool-output-dir",
        required=True,
        type=Path,
        help="Path to directories containing the verschillentool output.",
    )
    parser.add_argument(
        "--output-dir",
        required=True,
        type=Path,
        help="Path to directory where the output excel files will be written.",
    )
    return parser


if __name__ == "__main__":
    """Command line program to summarize the verschillentool output in a few excel files."""
    args = make_argument_parser().parse_args()

    verschillentool_output_dir: Path = args.verschillentool_output_dir
    if not verschillentool_output_dir.exists():
        raise NotADirectoryError(f"Directory '{verschillentool_output_dir}' does not exist.")

    output_dir: Path = args.output_dir
    output_dir.mkdir(parents=True, exist_ok=True)

    for file_type in FileType:
        statistics_per_model = parse_dataframes(verschillentool_output_dir, file_type=file_type)
        write_excel_flow_velocity(output_dir, statistics_per_model, file_type=file_type)
        write_excel_water_level(output_dir, statistics_per_model, file_type=file_type)
