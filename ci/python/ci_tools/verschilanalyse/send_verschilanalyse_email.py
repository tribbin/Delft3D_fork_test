import argparse
import html
import logging
import os
import smtplib
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import pandas as pd
from openpyxl import load_workbook

FOLDER_PATH = "./report"
ATTACHMENT_PATHS = [
    f"{FOLDER_PATH}/his_waterlevel.xlsx",
    f"{FOLDER_PATH}/his_flowvelocity.xlsx",
    f"{FOLDER_PATH}/map_waterlevel.xlsx",
    f"{FOLDER_PATH}/map_flowvelocity.xlsx",
]


def create_parser() -> argparse.ArgumentParser:
    """Make command line argument parser for the send_mail command."""
    parser = argparse.ArgumentParser(description="Send the weekly verschilanalyse email.")
    parser.add_argument("--build-id", required=True, help="The teamcity build identifier of this build.")
    parser.add_argument("--status", required=True, help="The status of the report processing step.")
    parser.add_argument("--teamcity-server-url", required=True, help="Base url for the teamcity server instance.")
    parser.add_argument("--build-type-id", required=True, help="Build ID that has the verschil analyse report.")
    parser.add_argument("--email-from", required=True, help="Email address of the sender.")
    parser.add_argument("--email-server", required=True, help="Adress of the email server.")
    parser.add_argument("--email-port", required=True, help="Port of the email server.")
    parser.add_argument("--email-recipients", required=True, help="List (csv) with email address recipients.")
    return parser


def get_attached_files_and_errors(
    attachment_paths: list[str],
) -> tuple[list[MIMEBase], list[str]]:
    """Attach files to the email message."""
    attachment_errors = []
    parts = []
    for attachment_path in attachment_paths:
        if os.path.exists(attachment_path):
            with open(attachment_path, "rb") as attachment:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(attachment.read())
                encoders.encode_base64(part)
                part.add_header(
                    "Content-Disposition",
                    f"attachment; filename={os.path.basename(attachment_path)}",
                )
                parts.append(part)
        else:
            attachment_errors.append(f"Attachment '{attachment_path}' not found.")
    return parts, attachment_errors


def create_email_message(email_from: str, email_recipients: str, email_content: str) -> MIMEMultipart:
    """Create the email message."""
    msg = MIMEMultipart()
    msg["Subject"] = "TeamCity VerschillenTool Build Report"
    msg["From"] = email_from
    msg["To"] = email_recipients
    msg.attach(MIMEText(email_content, "html"))
    return msg


def check_directories(path: str) -> str:
    """Check files in verschillentool output directories."""
    messages = []
    for root, dirs, _ in os.walk(path):
        for dir_name in dirs:
            dir_path = os.path.join(root, dir_name)
            files_in_dir = os.listdir(dir_path)
            if len(files_in_dir) == 1 and files_in_dir[0] == "log.txt":
                dir_path = dir_path[2:]
                messages.append(f"Something went wrong at '{dir_path}' as it only contains the file 'log.txt'.")
    if messages:
        return "\n".join(messages)
    else:
        return "OK"


def read_excel(file_path: str, sheet_name: str = "Averages") -> pd.DataFrame:
    """Read an Excel file and return the specified sheet as a DataFrame."""
    return pd.read_excel(file_path, sheet_name=sheet_name)


def check_for_red(file_paths: list[str], sheet_name: str = "Sheet") -> list[tuple[str, str]]:
    """Check for red cells in the specified Excel files and sheets."""
    red_rgb = "00FF0000"  # Define the RGB value for red

    rows: list[tuple[str, str]] = []

    for fp in file_paths:
        print(f"Checking file: {fp}")

        # Load the workbook with openpyxl to check formatting
        workbook = load_workbook(fp, data_only=True)
        sheet = workbook[sheet_name]

        red_cells = []

        # Iterate through rows and columns to find red cells
        for row in sheet.iter_rows():
            for cell in row:
                if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                    if cell.fill.fgColor.rgb == red_rgb:
                        red_cells.append(cell.coordinate)
                        rows.append((str(sheet[row[0].coordinate].value), fp))

    return sorted(rows)  # Tuples are sorted in lexicograhpical order.


def construct_email_content(
    file_paths: list[str],
    folder_path: str,
    hyperlink_build: str,
    hyperlink_logs: str,
    build_status: str,
) -> str:
    """Format the email contents using HTML."""
    rows = check_for_red(file_paths)
    if build_status == "success":
        check_message = check_directories(folder_path)
        if check_message == "OK":
            content = f'The build has been <span style="color:green;"> successful </span>: {hyperlink_build}'
        else:
            content = (
                f'The build has been <span style="color:green;"> successful </span> '
                f'and tolerances were <span style="color:green;"> not exceeded </span>, '
                f"but something(s) went wrong:\n\n"
                f"{check_message}\n\n"
                f"Link to build: {hyperlink_build}\n\n"
                f"Download the logs: {hyperlink_logs}\n\n"
            )
        if len(rows):
            content = "The build has been successful but tolerances were exceeded:\n\n"
            for model, excel in rows:
                content += f'Tolerance<span style="color:red;"> exceeded </span> for {model} in {excel}\n'
            content += f"\nLink to build: {hyperlink_build}\nDownload the logs: {hyperlink_logs}"
    elif build_status == "failure":
        content = f'The build has <span style="color:red;"> failed </span>: {hyperlink_build}'
    else:
        content = check_directories(folder_path)

    return content


def format_to_html(message: str, attachment_errors: list[str] | None = None) -> str:
    """Format HTML content of email."""
    formatted_message = message.replace("\n", "<br>")

    if attachment_errors:
        error_message = "<br>".join(attachment_errors)
        formatted_message += f"{error_message}"

    # Wrap the entire email content in basic HTML structure
    return f"""
    <html>
        <body>
            {formatted_message}
        </body>
    </html>
    """


def send_email(
    msg: MIMEMultipart,
    email_recipients: str,
    email_from: str,
    email_server: str,
    email_port: int,
) -> None:
    """Send verschilanalyse email."""
    try:
        with smtplib.SMTP(email_server, email_port) as server:
            server.sendmail(email_from, email_recipients, msg.as_string())
        logging.info("Email sent successfully.")
    except smtplib.SMTPException as e:
        logging.error(f"Failed to send email: {e}")
        raise


if __name__ == "__main__":
    """Command line program to send the weekly verschilanalyse email."""
    parser = create_parser()
    arguments = parser.parse_args()

    teamcity_server_url = arguments.teamcity_server_url.rstrip("/")
    build_type_url = f"{teamcity_server_url}/buildConfiguration/{arguments.build_type_id}"

    build_id: str = arguments.build_id
    build_status: str = arguments.status or "failure"
    if not build_id:
        build_status = "failure"

    build_url = f"{build_type_url}/{build_id}"
    link_text = "click here"
    hyperlink_build = f'<a href="{html.escape(build_url)}">{html.escape(link_text)}</a>'
    log_url = f"{teamcity_server_url}/repository/download/{arguments.build_type_id}/{build_id}:id/report.zip"
    hyperlink_logs = f'<a href="{html.escape(log_url)}">{html.escape(link_text)}</a>'

    email_content = construct_email_content(
        ATTACHMENT_PATHS, FOLDER_PATH, hyperlink_build, hyperlink_logs, build_status
    )

    attached_files, attachment_errors = get_attached_files_and_errors(ATTACHMENT_PATHS)

    if attachment_errors:
        html_email_content = format_to_html(email_content, attachment_errors)
        msg = create_email_message(arguments.email_from, arguments.email_recipients, html_email_content)
    else:
        html_email_content = format_to_html(email_content)
        msg = create_email_message(arguments.email_from, arguments.email_recipients, html_email_content)
        for attached_file in attached_files:
            msg.attach(attached_file)

    send_email(
        msg,
        arguments.email_recipients,
        arguments.email_from,
        arguments.email_server,
        int(arguments.email_port),
    )
