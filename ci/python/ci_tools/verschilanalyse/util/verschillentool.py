from dataclasses import dataclass
from enum import Enum
from typing import ClassVar

from openpyxl import Workbook


class OutputType(Enum):
    """The type of output file read by the verschillentool."""

    HIS = "his"
    MAP = "map"


class Variable(Enum):
    """The variable in the verschillentool output file.

    The unit of the water level is meters (m).
    The unit of the flow velocity is meters per second (m/s).
    """

    WATER_LEVEL = "water_level"  # in meters (m)
    FLOW_VELOCITY = "flow_velocity"  # in meters per second (m/s)

    @property
    def unit(self) -> str:
        """Get the unit of the variable."""
        return {
            Variable.WATER_LEVEL: "m",
            Variable.FLOW_VELOCITY: "m/s",
        }[self]


class Tolerances:
    """The maximum values of statistics above which a tolerance error is reported."""

    @staticmethod
    def max(output_type: OutputType, variable: Variable) -> float:
        """Get the tolerance for the `max` statistic given output type and variable.

        Parameters
        ----------
        output_type : OutputType
        variable : Variable

        Returns
        -------
        float
        """
        match output_type, variable:
            case OutputType.HIS, Variable.WATER_LEVEL:
                return HisWaterLevelTolerances.MAX
            case OutputType.HIS, Variable.FLOW_VELOCITY:
                return HisFlowVelocityTolerances.MAX
            case OutputType.MAP, Variable.WATER_LEVEL:
                return MapWaterLevelTolerances.MAX
            case OutputType.MAP, Variable.FLOW_VELOCITY:
                return MapFlowVelocityTolerances.MAX
            case _:
                raise ValueError(f"Unsupported output type {output_type} and variable {variable}")

    @staticmethod
    def bias(output_type: OutputType, variable: Variable) -> float:
        """Get the tolerance for the `bias` statistic given output type and variable.

        Parameters
        ----------
        output_type : OutputType
        variable : Variable

        Returns
        -------
        float
        """
        match output_type, variable:
            case OutputType.HIS, Variable.WATER_LEVEL:
                return HisWaterLevelTolerances.BIAS
            case OutputType.HIS, Variable.FLOW_VELOCITY:
                return HisFlowVelocityTolerances.BIAS
            case OutputType.MAP, Variable.WATER_LEVEL:
                return MapWaterLevelTolerances.BIAS
            case OutputType.MAP, Variable.FLOW_VELOCITY:
                return MapFlowVelocityTolerances.BIAS
            case _:
                raise ValueError(f"Unsupported output type {output_type} and variable {variable}")

    @staticmethod
    def rms(output_type: OutputType, variable: Variable) -> float:
        """Get the tolerance for the `rms` statistic given output type and variable.

        Parameters
        ----------
        output_type : OutputType
        variable : Variable

        Returns
        -------
        float
        """
        match output_type, variable:
            case OutputType.HIS, Variable.WATER_LEVEL:
                return HisWaterLevelTolerances.RMS
            case OutputType.HIS, Variable.FLOW_VELOCITY:
                return HisFlowVelocityTolerances.RMS
            case OutputType.MAP, Variable.WATER_LEVEL:
                return MapWaterLevelTolerances.RMS
            case OutputType.MAP, Variable.FLOW_VELOCITY:
                return MapFlowVelocityTolerances.RMS
            case _:
                raise ValueError(f"Unsupported output type {output_type} and variable {variable}")


class HisWaterLevelTolerances:
    """The maximum values of statistics above which a tolerance error is reported."""

    MAX: ClassVar[float] = 0.01  # m
    RMS: ClassVar[float] = 0.001  # m
    BIAS: ClassVar[float] = 0.0001  # m


class HisFlowVelocityTolerances:
    """The maximum values of statistics above which a tolerance error is reported."""

    MAX: ClassVar[float] = 0.05  # m/s
    RMS: ClassVar[float] = 0.005  # m/s
    BIAS: ClassVar[float] = 0.0005  # m/s


class MapWaterLevelTolerances:
    """The maximum values of statistics above which a tolerance error is reported."""

    MAX: ClassVar[float] = 0.05  # m
    RMS: ClassVar[float] = 0.001  # m
    BIAS: ClassVar[float] = 0.0001  # m


class MapFlowVelocityTolerances:
    """The maximum values of statistics above which a tolerance error is reported."""

    MAX: ClassVar[float] = 0.1  # m/s
    RMS: ClassVar[float] = 0.005  # m/s
    BIAS: ClassVar[float] = 0.0005  # m/s


@dataclass
class Statistics:
    """Contains statistics of a sample."""

    avg_max: float
    avg_bias: float
    avg_rms: float
    max: float


@dataclass
class VerschillentoolOutput:
    """Contains statistics of a model run."""

    output_type: OutputType
    flow_velocity: Statistics
    water_level: Statistics
    row_count: int

    @staticmethod
    def from_verschillentool_workbook(workbook: Workbook, output_type: OutputType) -> "VerschillentoolOutput":
        """Make a `VerschillentoolOutput` from a verschillentool excel file.

        This factory method produces an `VerschillentoolOuptut` from a `Workbook`,
        which is an excel file opened with `openpyxl`. The excel file is produced
        by the "verschillentool" and this class makes assumptions about the structure
        of the excel file (names of sheets, content of cells, etc...).

        Parameters
        ----------
        workbook : Workbook
            A workbook created from a verschillentool output excel file.
        output_type : OutputType
            The type of output file used to generate the excel file (map or his).

        Returns
        -------
        VerschillentoolOutput
            The information found in the verschillentool excel file.
        """
        averages_sheet = workbook["Averages"]
        statistics_sheet = workbook["Statistics"]
        maxima_sheet = workbook["Maxima"]

        stats_dict = {
            str(name_cell.value).split(maxsplit=1)[0]: float(value_cell.value)
            for name_cell, value_cell in averages_sheet["A2:B7"]
        }

        first_col = maxima_sheet.min_column - 1
        last_col = maxima_sheet.max_column - 2

        maxima_dict = {
            str(maxima_sheet[row][first_col].value).split(maxsplit=1)[0]: float(
                maxima_sheet[row][last_col].value  # type: ignore
            )
            for row in range(2, maxima_sheet.max_row + 1)
        }

        try:
            flow_velocity_stats = Statistics(
                avg_max=stats_dict["sea_water_speed_max"],
                avg_bias=stats_dict["sea_water_speed_bias"],
                avg_rms=stats_dict["sea_water_speed_rms"],
                max=maxima_dict["sea_water_speed"],
            )
            water_level_stats = Statistics(
                avg_max=stats_dict["sea_surface_height_max"],
                avg_bias=stats_dict["sea_surface_height_bias"],
                avg_rms=stats_dict["sea_surface_height_rms"],
                max=maxima_dict["sea_surface_height"],
            )
        except KeyError as exc:
            raise ValueError(f"Failed to parse verschillentool output: Missing key {exc}") from exc

        return VerschillentoolOutput(
            output_type=output_type,
            flow_velocity=flow_velocity_stats,
            water_level=water_level_stats,
            row_count=statistics_sheet.max_row - 1,  # The number of rows (minus one header row)
        )
