import itertools
from typing import ClassVar, Sequence

from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from ci_tools.verschilanalyse.util.slurm_log_data import LogComparison, SlurmLogData, Status
from ci_tools.verschilanalyse.util.verschilanalyse_comparison import VerschilanalyseComparison
from ci_tools.verschilanalyse.util.verschillentool import OutputType, Tolerances, Variable, VerschillentoolOutput


class ExcelExporter:
    """Contains code to generate the excel file attachment for the weekly verschilanalyse email."""

    RED: ClassVar[Color] = Color("FF0000")
    YELLOW: ClassVar[Color] = Color("FFFF00")
    GREEN: ClassVar[Color] = Color("00FF00")

    LOG_COMPARISON_SHEET_NAME = "Log comparisons"
    LOG_COMPARISON_DESCRIPTIONS: ClassVar[list[str]] = [
        "Commit ID",
        "Hostname",
        "Task count",
        "Mean computation time (s)",
        "Stddev computation time (s)",
        "WARNING line count",
        "ERROR line count",
        "Speedup",
        "Computation time difference (s)",
    ]

    VERSCHILLENTOOL_SHEET_NAMES = {
        OutputType.HIS: "His file statistics",
        OutputType.MAP: "Map file statistics",
    }
    VERSCHILLENTOOL_COUNT_HEADERS = {
        OutputType.HIS: "Station count",
        OutputType.MAP: "Map count",
    }

    @classmethod
    def _append_row(cls, sheet: Worksheet, model_name: str, stats: VerschillentoolOutput, ndigits: int = 4) -> None:
        sheet.append(
            [
                model_name,
                stats.row_count,
                round(stats.water_level.avg_max, ndigits=ndigits),
                round(stats.water_level.avg_mean, ndigits=ndigits),
                round(stats.water_level.avg_rms, ndigits=ndigits),
                round(stats.water_level.max, ndigits=ndigits),
                round(stats.flow_velocity.avg_max, ndigits=ndigits),
                round(stats.flow_velocity.avg_mean, ndigits=ndigits),
                round(stats.flow_velocity.avg_rms, ndigits=ndigits),
                round(stats.flow_velocity.max, ndigits=ndigits),
            ]
        )

        row = sheet[sheet.max_row]
        red_fill = cls._status_to_fill(Status.ERROR)
        # Apply the red style to cells based on the thresholds.
        if stats.water_level.avg_max > Tolerances.max(stats.output_type, Variable.WATER_LEVEL):
            row[2].fill = red_fill
            row[2].value = f"❌ {row[2].value}"
        if stats.water_level.avg_mean > Tolerances.mean(stats.output_type, Variable.WATER_LEVEL):
            row[3].fill = red_fill
            row[3].value = f"❌ {row[3].value}"
        if stats.water_level.avg_rms > Tolerances.rms(stats.output_type, Variable.WATER_LEVEL):
            row[4].fill = red_fill
            row[4].value = f"❌ {row[4].value}"
        if stats.flow_velocity.avg_max > Tolerances.max(stats.output_type, Variable.FLOW_VELOCITY):
            row[6].fill = red_fill
            row[6].value = f"❌ {row[6].value}"
        if stats.flow_velocity.avg_mean > Tolerances.mean(stats.output_type, Variable.FLOW_VELOCITY):
            row[7].fill = red_fill
            row[7].value = f"❌ {row[7].value}"
        if stats.flow_velocity.avg_rms > Tolerances.rms(stats.output_type, Variable.FLOW_VELOCITY):
            row[8].fill = red_fill
            row[8].value = f"❌ {row[8].value}"

    @staticmethod
    def _to_column(log_data: SlurmLogData | None) -> Sequence[str | int | float]:
        if log_data is None:
            return 7 * ["-"]

        crashed = log_data.is_crash()
        return [
            "-" if log_data.commit_id is None else log_data.commit_id,
            "-" if log_data.slurm_nodelist is None else log_data.slurm_nodelist,
            "-" if log_data.task_count == 0 else log_data.task_count,
            "-" if crashed else round(log_data.mean_computation_time, 3),
            "-" if crashed else round(log_data.stddev_computation_time, 3),
            len(log_data.warning_lines),
            len(log_data.error_lines),
        ]

    @classmethod
    def _status_to_fill(cls, status: Status) -> PatternFill:
        match status:
            case Status.ERROR:
                return PatternFill(fill_type="solid", start_color=cls.RED, end_color=cls.RED)
            case Status.WARNING:
                return PatternFill(fill_type="solid", start_color=cls.YELLOW, end_color=cls.YELLOW)
            case Status.SUCCESS:
                return PatternFill(fill_type="solid", start_color=cls.GREEN, end_color=cls.GREEN)
            case _:
                raise ValueError("Invalid icon")

    @classmethod
    def _make_log_comparison_sheet(
        cls,
        sheet: Worksheet,
        current_prefix: str,
        reference_prefix: str,
        log_comparisons: dict[str, LogComparison],
    ) -> None:
        # Add some information in this sheet about which output files we are comparing.
        for report_row in [
            ["Current verschilanalyse output", f"{current_prefix}/output"],
            ["Current logs", f"{current_prefix}/logs/logs.zip"],
            ["Reference verschilanalyse output", f"{reference_prefix}/output"],
            ["Reference logs", f"{reference_prefix}/logs/logs.zip"],
            [],
        ]:
            sheet.append(report_row)

        # Summarize the log comparisons.
        for model_name, comparison in sorted(log_comparisons.items()):
            current = comparison.current
            reference = comparison.reference
            speedup = comparison.speedup() or "-"
            time_diff = comparison.computation_time_difference() or "-"

            # Display 'error', 'warning', 'success' status of comparison.
            status_bar = comparison.get_comparison_status()

            # Write header with status.
            sheet.append([model_name, "Current", "Reference"])
            for i, status in enumerate(status_bar):
                cell = sheet[sheet.max_row][i]
                cell.fill = cls._status_to_fill(status)
                cell.value = f"{status.value} {cell.value}"

            description_column = cls.LOG_COMPARISON_DESCRIPTIONS
            current_column = itertools.chain(cls._to_column(current), [speedup, time_diff])
            reference_column = cls._to_column(reference)
            for row in itertools.zip_longest(description_column, current_column, reference_column, fillvalue=""):
                sheet.append(row)
            sheet.append([])

    @classmethod
    def _make_verschillentool_sheet(
        cls,
        sheet: Worksheet,
        output_type: OutputType,
        model_stats: dict[str, VerschillentoolOutput],
        ndigits: int = 4,
    ) -> None:
        if output_type == OutputType.HIS:
            unit = "stations"
        elif output_type == OutputType.MAP:
            unit = "times"
        else:
            raise ValueError(f"Invalid OutputType {output_type}")

        other_headers = [
            f"Maximum water level averaged over {unit} ({Variable.WATER_LEVEL.unit})",
            f"Bias water level averaged over {unit} ({Variable.WATER_LEVEL.unit})",
            f"RMSE water level averaged over {unit} ({Variable.WATER_LEVEL.unit})",
            f"Maximum water level over all {unit} ({Variable.WATER_LEVEL.unit})",
            f"Maximum flow velocity averaged over {unit} ({Variable.FLOW_VELOCITY.unit})",
            f"Bias flow velocity averaged over {unit} ({Variable.FLOW_VELOCITY.unit})",
            f"RMSE flow velocity averaged over {unit} ({Variable.FLOW_VELOCITY.unit})",
            f"Maximum flow velocity over all {unit} ({Variable.FLOW_VELOCITY.unit})",
        ]

        count_header = cls.VERSCHILLENTOOL_COUNT_HEADERS[output_type]

        sheet.append(["Model name", count_header, *other_headers])
        for model, stats in sorted(model_stats.items()):
            cls._append_row(sheet, model, stats, ndigits=ndigits)

    @classmethod
    def make_summary_workbook(cls, verschilanalyse: VerschilanalyseComparison, ndigits: int = 4) -> Workbook:
        """Make an excel workbook containing information found in this weekly verschilanalyse.

        The excel file contains more information than the email. It contains the
        statistics found in the verschillentool output excel files. It should be
        clear from the formatting which statistics exceeds the tolerances we have
        set for the weekly automated verschilanalyse.
        The excel file contains three sheets. One sheet summarizing the information
        found in the Slurm logs of each model run (from the current verschilanalyse
        and the reference verschilanalyse). The last two sheets contain the statistics
        reported by the verschillentool, for the `his` and `map` files respectively.

        Parameters
        ----------
        verschilanalyse : VerschilanalyseComparison
            An object containing all of the information collected for
            the weekly automated verschilanalyse.
        ndigits : int, optional
            Round the numbers in the excel file to `ndigits` amount of
            decimals after the decimal point. The default is four digits.

        Returns
        -------
        Workbook
            The excel workbook to write in the email attachment.
        """
        workbook = Workbook()
        log_comp_sheet = workbook.active
        if log_comp_sheet is None:
            raise RuntimeError("New workbook has no active sheet")
        log_comp_sheet.title = cls.LOG_COMPARISON_SHEET_NAME

        log_comparisons = verschilanalyse.get_log_comparisons()
        cls._make_log_comparison_sheet(
            sheet=log_comp_sheet,
            current_prefix=verschilanalyse.s3_current_prefix,
            reference_prefix=verschilanalyse.s3_reference_prefix,
            log_comparisons=log_comparisons,
        )

        his_sheet: Worksheet = workbook.create_sheet(title=cls.VERSCHILLENTOOL_SHEET_NAMES[OutputType.HIS])
        cls._make_verschillentool_sheet(his_sheet, OutputType.HIS, verschilanalyse.his_outputs, ndigits=ndigits)

        map_sheet: Worksheet = workbook.create_sheet(title=cls.VERSCHILLENTOOL_SHEET_NAMES[OutputType.MAP])
        cls._make_verschillentool_sheet(map_sheet, OutputType.MAP, verschilanalyse.map_outputs, ndigits=ndigits)

        return workbook
